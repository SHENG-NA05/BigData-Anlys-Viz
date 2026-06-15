import csv
import re
from pathlib import Path
from typing import BinaryIO, TextIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.crud.catalog import create_catalog_books
from app.db.models import CatalogBook


REQUIRED_COLUMNS = {"title", "isbn", "classification_no"}
OPTIONAL_COLUMNS = {
    "author",
    "publisher",
    "publication_year",
    "summary",
}
SUPPORTED_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS
CLASSIFICATION_NO_PATTERN = re.compile(r"^\d{3}(?:\.\d{1,3})?$")


class CatalogImportError(ValueError):
    pass


class CatalogService:
    def import_file_upload(self, db: Session, upload_file: BinaryIO, filename: str) -> int:
        lower_filename = filename.lower()
        if lower_filename.endswith(".csv"):
            return self.import_csv_upload(db, upload_file, filename)
        if lower_filename.endswith((".xlsx", ".xls")):
            return self.import_excel_upload(db, upload_file, filename)
        raise CatalogImportError("Only CSV and Excel files are supported")

    def import_csv_path(self, db: Session, csv_path: str | Path) -> int:
        path = Path(csv_path)
        if not path.exists():
            raise CatalogImportError(f"CSV file not found: {path}")

        with path.open("r", encoding="utf-8-sig", newline="") as file:
            return self.import_csv(db, file, source_file=path.name)

    def import_excel_path(self, db: Session, excel_path: str | Path) -> int:
        path = Path(excel_path)
        if not path.exists():
            raise CatalogImportError(f"Excel file not found: {path}")

        with path.open("rb") as file:
            return self.import_excel_upload(db, file, path.name)

    def import_csv_upload(self, db: Session, upload_file: BinaryIO, filename: str) -> int:
        text_file = _decode_upload_file(upload_file)
        return self.import_csv(db, text_file, source_file=filename)

    def import_excel_upload(self, db: Session, upload_file: BinaryIO, source_file: str) -> int:
        rows = _read_excel_rows(upload_file)
        return self.import_rows(db, rows, source_file=source_file, empty_message="Excel file has no catalog rows")

    def import_csv(self, db: Session, csv_file: TextIO, source_file: str | None = None) -> int:
        reader = csv.DictReader(csv_file)
        self._validate_headers(reader.fieldnames)
        return self.import_rows(
            db,
            list(reader),
            source_file=source_file,
            start_row_number=2,
            empty_message="CSV file has no catalog rows",
        )

    def import_rows(
        self,
        db: Session,
        rows: list[dict[str, object]],
        source_file: str | None = None,
        start_row_number: int = 2,
        empty_message: str = "File has no catalog rows",
    ) -> int:
        books = []
        errors = []
        for row_number, row in enumerate(rows, start=start_row_number):
            try:
                books.append(self._row_to_book(row, source_file))
            except CatalogImportError as exc:
                errors.append(f"row {row_number}: {exc}")

        if errors:
            raise CatalogImportError("; ".join(errors[:10]))

        if not books:
            raise CatalogImportError(empty_message)

        return create_catalog_books(db, books)

    def _validate_headers(self, fieldnames: list[str] | None) -> None:
        if not fieldnames:
            raise CatalogImportError("CSV file has no header row")

        normalized = {_normalize_header(name) for name in fieldnames}
        missing = REQUIRED_COLUMNS - normalized
        if missing:
            raise CatalogImportError(f"Missing required columns: {', '.join(sorted(missing))}")

    def _row_to_book(self, row: dict[str, object], source_file: str | None) -> CatalogBook:
        normalized_row = {_normalize_header(key): _clean_value(value) for key, value in row.items()}

        title = normalized_row.get("title")
        isbn = normalized_row.get("isbn")
        classification_no = normalized_row.get("classification_no")
        if not title:
            raise CatalogImportError("title is required")
        if not isbn:
            raise CatalogImportError("isbn is required")
        if not classification_no:
            raise CatalogImportError("classification_no is required")
        _validate_classification_no(classification_no)

        publication_year = _parse_publication_year(normalized_row.get("publication_year"))

        return CatalogBook(
            title=title,
            isbn=isbn,
            author=normalized_row.get("author"),
            publisher=normalized_row.get("publisher"),
            publication_year=publication_year,
            classification_no=classification_no,
            summary=normalized_row.get("summary"),
            source_file=source_file,
        )


def _decode_upload_file(upload_file: BinaryIO) -> TextIO:
    import io

    raw = upload_file.read()
    if isinstance(raw, str):
        return io.StringIO(raw)
    return io.StringIO(raw.decode("utf-8-sig"))


def _read_excel_rows(upload_file: BinaryIO) -> list[dict[str, object]]:
    workbook = load_workbook(upload_file, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise CatalogImportError("Excel file has no header row") from exc

    fieldnames = [_clean_value(value) for value in header_row]
    if not any(fieldnames):
        raise CatalogImportError("Excel file has no header row")
    CatalogService()._validate_headers([str(name) if name is not None else "" for name in fieldnames])

    records = []
    for row in rows:
        if not any(_clean_value(value) for value in row):
            continue
        records.append(
            {
                str(header): value
                for header, value in zip(fieldnames, row)
                if header is not None
            }
        )
    return records


def _normalize_header(value: str | None) -> str:
    return str(value or "").strip().lower()


def _clean_value(value: object | None) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _parse_publication_year(value: str | None) -> int | None:
    if not value:
        return None
    try:
        year = int(value)
    except ValueError as exc:
        raise CatalogImportError(f"publication_year must be an integer: {value}") from exc
    if year < 0:
        raise CatalogImportError(f"publication_year must be positive: {value}")
    return year


def _validate_classification_no(value: str) -> None:
    if not CLASSIFICATION_NO_PATTERN.match(value):
        raise CatalogImportError(
            "classification_no must be 3 digits with optional decimal digits, "
            f"for example 540 or 540.123: {value}"
        )
